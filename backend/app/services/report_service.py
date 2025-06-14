from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Any
import io
import os

class ReportService:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#2563eb')
        )
    
    def generate_employee_performance_pdf(self, employee_data: Dict, performance_data: List[Dict]) -> bytes:
        """Generate employee performance report in PDF format"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        title = Paragraph(f"Employee Performance Report - {employee_data['full_name']}", self.title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Employee Info
        info_data = [
            ['Employee ID:', employee_data['employee_id']],
            ['Department:', employee_data['department']],
            ['Designation:', employee_data['designation']],
            ['Report Period:', f"{datetime.now().strftime('%B %Y')}"],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Performance Metrics
        if performance_data:
            perf_title = Paragraph("Performance Metrics", self.styles['Heading2'])
            story.append(perf_title)
            story.append(Spacer(1, 12))
            
            perf_headers = ['Metric', 'Score', 'Rating']
            perf_rows = [perf_headers]
            
            for perf in performance_data:
                rating = self._get_performance_rating(perf['overall_score'])
                perf_rows.append([
                    'Overall Performance',
                    f"{perf['overall_score']}/10",
                    rating
                ])
                perf_rows.append([
                    'Technical Skills',
                    f"{perf['technical_skills']}/10",
                    self._get_performance_rating(perf['technical_skills'])
                ])
                perf_rows.append([
                    'Communication',
                    f"{perf['communication']}/10",
                    self._get_performance_rating(perf['communication'])
                ])
            
            perf_table = Table(perf_rows, colWidths=[2*inch, 1*inch, 1.5*inch])
            perf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(perf_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_department_report_excel(self, department: str, metrics_data: List[Dict]) -> bytes:
        """Generate department report in Excel format"""
        buffer = io.BytesIO()
        workbook = Workbook()
        
        # Main sheet
        ws = workbook.active
        ws.title = f"{department} Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        
        # Title
        ws['A1'] = f"{department} Department Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:F2')
        
        # Headers
        headers = ['Metric', 'Current Month', 'Previous Month', 'Change', 'Target', 'Achievement %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, data in enumerate(metrics_data, 5):
            ws.cell(row=row, column=1, value=data.get('metric', ''))
            ws.cell(row=row, column=2, value=data.get('current_value', 0))
            ws.cell(row=row, column=3, value=data.get('previous_value', 0))
            ws.cell(row=row, column=4, value=data.get('change_percent', 0))
            ws.cell(row=row, column=5, value=data.get('target', 0))
            ws.cell(row=row, column=6, value=data.get('achievement_percent', 0))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_company_analytics_pdf(self, analytics_data: Dict) -> bytes:
        """Generate company-wide analytics report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        title = Paragraph("Company Analytics Report", self.title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Executive Summary
        summary_title = Paragraph("Executive Summary", self.styles['Heading2'])
        story.append(summary_title)
        story.append(Spacer(1, 12))
        
        summary_data = [
            ['Total Revenue', f"${analytics_data.get('total_revenue', 0):,.2f}"],
            ['Net Profit', f"${analytics_data.get('net_profit', 0):,.2f}"],
            ['Total Employees', str(analytics_data.get('total_employees', 0))],
            ['Active Projects', str(analytics_data.get('active_projects', 0))],
            ['Customer Satisfaction', f"{analytics_data.get('customer_satisfaction', 0):.1f}%"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Department Performance
        dept_title = Paragraph("Department Performance", self.styles['Heading2'])
        story.append(dept_title)
        story.append(Spacer(1, 12))
        
        dept_data = analytics_data.get('department_performance', [])
        if dept_data:
            dept_headers = ['Department', 'Revenue Contribution', 'Employee Count', 'Productivity Score']
            dept_rows = [dept_headers]
            
            for dept in dept_data:
                dept_rows.append([
                    dept.get('department', ''),
                    f"${dept.get('revenue_contribution', 0):,.2f}",
                    str(dept.get('employee_count', 0)),
                    f"{dept.get('productivity_score', 0):.1f}%"
                ])
            
            dept_table = Table(dept_rows, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1.5*inch])
            dept_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(dept_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _get_performance_rating(self, score: float) -> str:
        """Convert numeric score to performance rating"""
        if score >= 9:
            return "Excellent"
        elif score >= 7:
            return "Good"
        elif score >= 5:
            return "Average"
        elif score >= 3:
            return "Below Average"
        else:
            return "Poor"

# Global instance
report_service = ReportService()